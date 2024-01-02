#!/usr/bin/env python
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from datetime import datetime
from email.utils import formatdate

# Function to send an email
def send_email(receiver_email, subject, body):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain"))

    # Establish a connection with the SMTP server
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, receiver_email, message.as_string())

# Replace these with your Gmail credentials
sender_email = "trailidsam@gmail.com"
sender_password = "sufapdhwpmytxyla"

# Replace 'your_file.xlsx' with the actual file name
file_path = r'C:\Users\Sampath\Documents\details.xlsx'

# Load the workbook
workbook = load_workbook(file_path)

# Select the active sheet (you can change 'Sheet1' to the sheet name you want)
sheet = workbook['Sheet1']

# Get today's date in the format of your Excel sheet's DOB column
today_date = datetime.now().strftime("%m-%d")

# Initialize lists for today's birthdays and others
today_birthdays = []
other_birthdays = []

# Iterate through rows
for row in sheet.iter_rows(values_only=True):
    if len(row) >= 3:
        name, dob_str, gmail_id = row[:3]
        # Assume dob_str is in the format "%m-%d"
        dob_str = f"{datetime.now().year}-{dob_str}"
        dob = datetime.strptime(dob_str, "%Y-%m-%d")
        # Check if the person's birthday is today
        if dob.strftime("%m-%d") == today_date:
            today_birthdays.append((name, dob, gmail_id))
        else:
            other_birthdays.append((name, dob, gmail_id))

# Close the workbook when done
workbook.close()

# Send emails based on the number of birthdays today
if len(today_birthdays) == 1:
    # If only one person has a birthday today
    name, dob, receiver_email = today_birthdays[0]
    birthday_message = f"""
Hey {name}!

A2B2 wishes you a fantastic and joy-filled birthday! 🎉🥳 Wishing you a day filled with laughter, joy, and all the things that make you smile! 🥳 May this year bring you incredible adventures, endless happiness, and dreams come true.

Cheers to another fantastic journey around the sun! 🌞🎈

Happy Birthday {name}!

Warmest wishes,
Team VANQUISHERS !
"""
    # Uncomment the line below to send the email
    send_email(receiver_email, f"Happy Birthday {name}!", birthday_message)
    print(f"Sent email to {receiver_email}")

else:
    # If more than one person has a birthday today
    for person in today_birthdays:
        name, dob, receiver_email = person
        other_persons = [p[0] for p in today_birthdays if p != person]
        birthday_message = f"""
Hey {name}!

A2B2 wishes you a fantastic and joy-filled birthday! 🎉🥳 Wishing you a day filled with laughter, joy, and all the things that make you smile! 🥳 May this year bring you incredible adventures, endless happiness, and dreams come true.

Cheers to another fantastic journey around the sun! 🌞🎈

Happy Birthday {name}!

Warmest wishes,
Team VANQUISHERS !

Today is also the birthday of {', '.join(other_persons)}.Feel free to send your warm wishes and make their day extra special.

"""
        
        # Uncomment the line below to send the email
        send_email(receiver_email, f"Happy Birthday {name}!", birthday_message)
        print(f"Sent email to {receiver_email}")

# Sending reminder emails for other birthdays
# Sending reminder emails for other birthdays
for person in other_birthdays:
    name, dob, receiver_email = person
    reminder_message = f"""
Hi {name}!

Just a quick note to make sure you don't miss out on the celebration today! Today is the birthday of {', '.join([p[0] for p in today_birthdays])}, and I thought you'd like to join in the festivities. 🥳🎈

Feel free to send your warm wishes and make their day extra special. Let's add some more joy to the occasion!

Thanks for sharing in the excitement!

Best,
Team VANQUIHERS !
"""
    send_email(receiver_email, "Reminder Mail", reminder_message)
    print(f"Sent email to {receiver_email}")

